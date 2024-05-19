import os
import re
import sys
import shutil
import gspread
import warnings
import threading
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from gspread_dataframe import set_with_dataframe
from google.oauth2.service_account import Credentials

SCOPES = ['https://www.googleapis.com/auth/drive', 'https://www.googleapis.com/auth/spreadsheets']

# Path to the service account key file
XLFolder = "xlFile.txt"
SERVICE_FILE = 'service_account.json'

def gc_creator():
    """
    This function selects a service account credentials file. If the file
    does not exist in the current directory, it prompts the user to select
    the file via a file dialog and then copies it to the current directory.
    """
    if os.path.exists(SERVICE_FILE):
        credentials = Credentials.from_service_account_file(SERVICE_FILE, scopes=SCOPES)
        gc = gspread.authorize(credentials)
    else:
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        file_path = filedialog.askopenfilename(
            title="Select credentials.json file",
            filetypes=[("JSON files", "*.json")]
        )

        if file_path:
            # Copy the file to the current directory
            new_file_path = os.path.join(os.getcwd(), SERVICE_FILE)
            shutil.copyfile(file_path, new_file_path)
            credentials = Credentials.from_service_account_file(SERVICE_FILE, scopes=SCOPES)
            gc = gspread.authorize(credentials)
        else:
            raise FileNotFoundError("No file selected for credentials.")
    return gc

def load_xl_folder():
    """
    This function loads the path of the folder containing Excel files.
    If the path is not saved, it prompts the user to select the folder.
    """
    if os.path.exists(XLFolder):
        with open(XLFolder, 'r') as f:
            return f.read().strip()
    else:
        return xlfileloader()

def xlfileloader():
    """
    This function prompts the user to select a folder and saves its path.
    """
    try:
        root = tk.Tk()
        root.withdraw()  # Hide the main window

        # Ask the user to select a directory
        folder_path = filedialog.askdirectory(
            title="Select Task Folder"
        )
        if folder_path:
            with open(XLFolder, 'w') as f:
                f.write(folder_path)
        return folder_path if folder_path else None
    except Exception as e:
        print(f"An error occurred while selecting the folder: {e}")
        return None
    finally:
        root.destroy()

def excel_cell_to_indices(cell):
    """
    This function converts an Excel cell reference (e.g., 'B2') to zero-based row and column indices.
    """
    match = re.match(r"([A-Z]+)([0-9]+)", cell.upper())
    if not match:
        raise ValueError(f"Invalid Excel cell reference: {cell}")
    column = match.group(1)
    row = int(match.group(2)) - 1  # Excel rows are 1-based, convert to 0-based
    column_number = 0
    for char in column:
        column_number = column_number * 26 + (ord(char) - ord('A')) + 1
    return row, column_number  # Convert to zero-based index

def main():
    xlfolder = load_xl_folder()
    if not xlfolder:
        print("No folder selected. Exiting...")
        return

    task_file = os.path.join(xlfolder, 'Task File.xlsx')
    try:
        task_df = pd.read_excel(task_file, header=1)
        task_df = task_df.dropna(how='all')
    except Exception as e:
        print(f"Error reading the task file: {e}")
        return

    xl_sheet = task_df["Source Excell Spreadsheet"].tolist()
    xl_start_cell = task_df["Starting CELL"].tolist()
    xl_end_cell = task_df["Ending CELL * means all"].tolist()
    gl_sheetID = task_df["Google Sheet ID"].tolist()
    gl_sheetTab = task_df["Tab"].tolist()
    gl_startCell = task_df["Starting Cell"].tolist()
    gl_Duplicate = task_df["Duplicate Tab"].tolist()

    client = gc_creator()

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        for i, sheet_name in enumerate(xl_sheet):
            try:
                xl_file = os.path.join(xlfolder, sheet_name)
                wb = load_workbook(xl_file, data_only=True)
                sheet = wb.active
                row_count = sheet.max_row
                column_count = sheet.max_column
                start_row, start_col = excel_cell_to_indices(xl_start_cell[i])
                
                if xl_end_cell[i] == '*':
                    end_row, end_col = row_count, column_count
                else:
                    end_row, end_col = excel_cell_to_indices(xl_end_cell[i])

                nrows = end_row - start_row + 1

                df = pd.read_excel(xl_file, skiprows=start_row-1, nrows=nrows)
                
                spreadsheet = client.open_by_key(gl_sheetID[i])

                if gl_sheetTab[i].lower() == 'create new':
                    new_sheet_name = 'Copy of ' + gl_Duplicate[i]
                    existing_sheets = spreadsheet.worksheets()
                    existing_sheet_names = [ws.title for ws in existing_sheets]

                # Check if the sheet already exists and delete it if necessary
                    if new_sheet_name in existing_sheet_names:
                        worksheet_to_delete = spreadsheet.worksheet(new_sheet_name)
                        spreadsheet.del_worksheet(worksheet_to_delete)

                    dupe_worksheet = spreadsheet.worksheet(gl_Duplicate[i])
                    worksheet = spreadsheet.duplicate_sheet(dupe_worksheet.id, new_sheet_name=new_sheet_name)
                    
                else:
                    worksheet = spreadsheet.worksheet(gl_sheetTab[i])
                    
                row_index, col_index = excel_cell_to_indices(gl_startCell[i])
                set_with_dataframe(worksheet, df, row=row_index+1, col=col_index, include_index=False, include_column_header=False)

                print(f"Data from '{sheet_name}' copied to {gl_sheetTab[i]}")
            except Exception as e:
                print(f"An error occurred while processing '{sheet_name}': {e}")


def run_main():
    root = tk.Tk()
    root.title("My App")
    root.geometry("400x400")  # Set the initial size of the window

    # Create a frame
    frame = tk.Frame(root)
    frame.pack(pady=20)

    # Create a scrollbar
    scrollbar = tk.Scrollbar(frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # Create the Text widget to display output
    output_text = tk.Text(frame, height=8, width=50, wrap="word", yscrollcommand=scrollbar.set)
    output_text.pack(pady=10)

    # Configure the scrollbar
    scrollbar.config(command=output_text.yview)

    # Create the Label widget to indicate completion
    completion_label = tk.Label(frame, text="", font=("Arial", 12, "bold"), pady=10)
    completion_label.pack()

    # Redirect the standard output to the Text widget
    sys.stdout = PrintToTextWidget(output_text)

    # Function to run main() and update completion label
    def run_code():
        completion_label.config(text="Running...")
        threading.Thread(target=run_main_thread).start()

    def run_main_thread():
        main()
        completion_label.config(text="Completed!")
        completion_label.pack(pady=10)
        
    def run_close():
        root.destroy()

    # Create the "Run" button
    run_button = tk.Button(frame, text="Run the Code", height=1, width=20, command=run_code)
    run_button.pack(pady=10)

    exit_button = tk.Button(frame, text="Exit Code", height=1, width=20, command=run_close)
    exit_button.pack(pady=10)
    
    # Run the Tkinter event loop
    root.mainloop()

class PrintToTextWidget:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, message):
        self.text_widget.insert(tk.END, message)
        self.text_widget.see(tk.END)

    def flush(self):
        pass

if __name__ == "__main__":
    run_main()