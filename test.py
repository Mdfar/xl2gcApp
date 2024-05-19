import os
import warnings
import pandas as pd
import tkinter as tk
import shutil
from tkinter import filedialog
import gspread
from google.oauth2.service_account import Credentials
from gspread_dataframe import set_with_dataframe
from openpyxl import load_workbook
import re
import sys
import threading

SCOPES = ['https://www.googleapis.com/auth/drive', 'https://www.googleapis.com/auth/spreadsheets']

# Path to the service account key file
XLFolder = "xlFile.txt"
SERVICE_FILE = 'service_account.json'

def gc_creator():
    """
    This function selects a service account credentials file. If the file
    does not exist in the current directory, it prompts the user to select
    the file via a file dialog and then copies it to the current directory.
    """
    if os.path.exists(SERVICE_FILE):
        credentials = Credentials.from_service_account_file(SERVICE_FILE, scopes=SCOPES)
        gc = gspread.authorize(credentials)
    else:
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        file_path = filedialog.askopenfilename(
            title="Select credentials.json file",
            filetypes=[("JSON files", "*.json")]
        )

        if file_path:
            # Copy the file to the current directory
            new_file_path = os.path.join(os.getcwd(), SERVICE_FILE)
            shutil.copyfile(file_path, new_file_path)
            credentials = Credentials.from_service_account_file(SERVICE_FILE, scopes=SCOPES)
            gc = gspread.authorize(credentials)
        else:
            raise FileNotFoundError("No file selected for credentials.")
    return gc

def load_xl_folder():
    """
    This function loads the path of the folder containing Excel files.
    If the path is not saved, it prompts the user to select the folder.
    """
    if os.path.exists(XLFolder):
        with open(XLFolder, 'r') as f:
            return f.read().strip()
    else:
        return xlfileloader()

def xlfileloader():
    """
    This function prompts the user to select a folder and saves its path.
    """
    try:
        root = tk.Tk()
        root.withdraw()  # Hide the main window

        # Ask the user to select a directory
        folder_path = filedialog.askdirectory(
            title="Select Task Folder"
        )
        if folder_path:
            with open(XLFolder, 'w') as f:
                f.write(folder_path)
        return folder_path if folder_path else None
    except Exception as e:
        print(f"An error occurred while selecting the folder: {e}")
        return None
    finally:
        root.destroy()

def excel_cell_to_indices(cell):
    """
    This function converts an Excel cell reference (e.g., 'B2') to zero-based row and column indices.
    """
    match = re.match(r"([A-Z]+)([0-9]+)", cell.upper())
    if not match:
        raise ValueError(f"Invalid Excel cell reference: {cell}")
    column = match.group(1)
    row = int(match.group(2)) -1  # Excel rows are 1-based, convert to 0-based
    column_number = 0
    for char in column:
        column_number = column_number * 26 + (ord(char) - ord('A')) + 1
    return row, column_number  # Convert to zero-based index

def main():
    xlfolder = load_xl_folder()
    if not xlfolder:
        print("No folder selected. Exiting...")
        return

    task_file = os.path.join(xlfolder, 'Task File.xlsx')
    try:
        task_df = pd.read_excel(task_file, header=1)
        task_df = task_df.dropna(how='all')
    except Exception as e:
        print(f"Error reading the task file: {e}")
        return

    xl_sheet = task_df["Source Excell Spreadsheet"].tolist()
    xl_start_cell = task_df["Starting CELL"].tolist()
    xl_end_cell = task_df["Ending CELL * means all"].tolist()
    gl_sheetID = task_df["Google Sheet ID"].tolist()
    gl_sheetTab = task_df["Tab"].tolist()
    gl_startCell = task_df["Starting Cell"].tolist()
    gl_Duplicate = task_df["Duplicate Tab"].tolist()

    client = gc_creator()

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        for i, sheet_name in enumerate(xl_sheet):
            xl_file = os.path.join(xlfolder, sheet_name)
            wb = load_workbook(xl_file, data_only=True)
            sheet = wb.active
            row_count = sheet.max_row
            column_count = sheet.max_column
            start_row, start_col = excel_cell_to_indices(xl_start_cell[i])
            
            if xl_end_cell[i] == '*':
                end_row, end_col = row_count, column_count
            else:
                end_row, end_col = excel_cell_to_indices(xl_end_cell[i])

            nrows = end_row - start_row + 1

            df = pd.read_excel(xl_file, skiprows=start_row-1, nrows=nrows)
            
            print(df.head(5))
if __name__ == '__main__':
    main()